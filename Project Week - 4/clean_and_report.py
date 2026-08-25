"""
Data Cleaning & Reporting Automation
=====================================
Internship task pipeline: reads a raw dataset, cleans it (missing values,
duplicates, inconsistent formatting), then automatically generates:
  1. A cleaned CSV
  2. A summary statistics table
  3. Chart images (visual summaries)
  4. A formatted Excel report combining tables + charts

Usage:
    python clean_and_report.py

To use with your own data, just change RAW_DATA_PATH below to point at your
file (CSV or Excel both work).
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
RAW_DATA_PATH = Path("data/raw_sales_data.csv")
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

CLEANED_CSV_PATH = OUTPUT_DIR / "cleaned_sales_data.csv"
CLEANING_LOG_PATH = OUTPUT_DIR / "cleaning_log.txt"
CHART_DIR = OUTPUT_DIR / "charts"
CHART_DIR.mkdir(exist_ok=True)
REPORT_PATH = OUTPUT_DIR / "sales_report.xlsx"


def log(message, log_lines):
    """Print and record a cleaning-step message."""
    print(message)
    log_lines.append(message)


# ---------------------------------------------------------------------------
# STEP 1: LOAD
# ---------------------------------------------------------------------------
def load_data(path):
    if path.suffix == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)


# ---------------------------------------------------------------------------
# STEP 2: CLEAN
# ---------------------------------------------------------------------------
def clean_data(df, log_lines):
    df = df.copy()
    log(f"Raw dataset shape: {df.shape}", log_lines)

    # --- Standardize text columns (whitespace, casing) ---
    text_cols = df.select_dtypes(include=["object", "string"]).columns
    for col in text_cols:
        if col == "Date":
            continue
        is_missing = df[col].isna()
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].str.replace(r"\s+", " ", regex=True)
        df[col] = df[col].str.title()
        df.loc[is_missing, col] = np.nan
    log(f"Standardized text formatting in columns: {list(text_cols)}", log_lines)

    # --- Parse inconsistent date formats ---
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", format="mixed")
        n_bad_dates = df["Date"].isna().sum()
        log(f"Parsed 'Date' column into a consistent datetime format "
            f"({n_bad_dates} unparseable dates set to NaT)", log_lines)

    # --- Remove exact duplicate rows ---
    n_before = len(df)
    df = df.drop_duplicates()
    log(f"Removed {n_before - len(df)} exact duplicate rows", log_lines)

    # --- Handle invalid values (e.g. negative units sold) ---
    if "UnitsSold" in df.columns:
        n_invalid = (df["UnitsSold"] < 0).sum()
        df.loc[df["UnitsSold"] < 0, "UnitsSold"] = np.nan
        log(f"Flagged {n_invalid} negative 'UnitsSold' values as invalid -> missing", log_lines)

    # --- Handle missing values ---
    missing_before = df.isna().sum()
    numeric_cols = df.select_dtypes(include=np.number).columns
    for col in numeric_cols:
        median_val = df[col].median()
        df[col] = df[col].fillna(median_val)
    for col in ["Region", "SalesRep", "Product"]:
        if col in df.columns:
            mode_val = df[col].mode(dropna=True)
            fill_val = mode_val.iloc[0] if not mode_val.empty else "Unknown"
            df[col] = df[col].fillna(fill_val)
    log("Filled missing numeric values with column median; "
        "missing categorical values with column mode:", log_lines)
    for col, n_missing in missing_before.items():
        if n_missing > 0:
            log(f"    - {col}: {n_missing} missing values filled", log_lines)

    # --- Recompute derived column ---
    if {"UnitsSold", "UnitPrice"}.issubset(df.columns):
        df["TotalSale"] = (df["UnitsSold"] * df["UnitPrice"]).round(2)

    df = df.reset_index(drop=True)
    log(f"Cleaned dataset shape: {df.shape}", log_lines)
    return df


# ---------------------------------------------------------------------------
# STEP 3: SUMMARY STATISTICS
# ---------------------------------------------------------------------------
def build_summary(df):
    summary = {
        "Total Orders": len(df),
        "Total Revenue": round(df["TotalSale"].sum(), 2),
        "Average Order Value": round(df["TotalSale"].mean(), 2),
        "Total Units Sold": int(df["UnitsSold"].sum()),
        "Date Range": f"{df['Date'].min().date()} to {df['Date'].max().date()}",
    }
    by_region = df.groupby("Region")["TotalSale"].sum().sort_values(ascending=False)
    by_product = df.groupby("Product")["TotalSale"].sum().sort_values(ascending=False)
    by_rep = df.groupby("SalesRep")["TotalSale"].sum().sort_values(ascending=False)
    monthly = df.set_index("Date").resample("ME")["TotalSale"].sum()
    return summary, by_region, by_product, by_rep, monthly


# ---------------------------------------------------------------------------
# STEP 4: VISUAL SUMMARIES
# ---------------------------------------------------------------------------
def make_charts(by_region, by_product, by_rep, monthly):
    plt.style.use("seaborn-v0_8-whitegrid")
    chart_paths = {}

    fig, ax = plt.subplots(figsize=(6, 4))
    by_region.plot(kind="bar", ax=ax, color="#4C72B0")
    ax.set_title("Revenue by Region")
    ax.set_ylabel("Total Sale ($)")
    fig.tight_layout()
    p = CHART_DIR / "revenue_by_region.png"
    fig.savefig(p, dpi=150)
    plt.close(fig)
    chart_paths["region"] = p

    fig, ax = plt.subplots(figsize=(6, 4))
    by_product.plot(kind="bar", ax=ax, color="#55A868")
    ax.set_title("Revenue by Product")
    ax.set_ylabel("Total Sale ($)")
    fig.tight_layout()
    p = CHART_DIR / "revenue_by_product.png"
    fig.savefig(p, dpi=150)
    plt.close(fig)
    chart_paths["product"] = p

    fig, ax = plt.subplots(figsize=(6, 4))
    by_rep.plot(kind="bar", ax=ax, color="#C44E52")
    ax.set_title("Revenue by Sales Rep")
    ax.set_ylabel("Total Sale ($)")
    fig.tight_layout()
    p = CHART_DIR / "revenue_by_rep.png"
    fig.savefig(p, dpi=150)
    plt.close(fig)
    chart_paths["rep"] = p

    fig, ax = plt.subplots(figsize=(6, 4))
    monthly.plot(kind="line", marker="o", ax=ax, color="#8172B2")
    ax.set_title("Monthly Revenue Trend")
    ax.set_ylabel("Total Sale ($)")
    fig.tight_layout()
    p = CHART_DIR / "monthly_trend.png"
    fig.savefig(p, dpi=150)
    plt.close(fig)
    chart_paths["trend"] = p

    return chart_paths


# ---------------------------------------------------------------------------
# STEP 5: EXCEL REPORT
# ---------------------------------------------------------------------------
def build_excel_report(df, summary, by_region, by_product, by_rep, chart_paths):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4C72B0", end_color="4C72B0", fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=14)
    body_font = Font(name="Arial", size=11)

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Automated Sales Data Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9)

    row = 4
    for label, value in summary.items():
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=2, value=value).font = body_font
        row += 1

    row += 1
    img_row = row
    for key in ["region", "product", "rep", "trend"]:
        img = XLImage(str(chart_paths[key]))
        img.width, img.height = 480, 320
        ws.add_image(img, f"A{img_row}")
        img_row += 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # --- Breakdown sheet ---
    ws2 = wb.create_sheet("Breakdowns")
    ws2["A1"] = "Revenue by Region"
    ws2["A1"].font = title_font
    _write_series(ws2, by_region, start_row=2, header_font=header_font, header_fill=header_fill, body_font=body_font)

    start = 2 + len(by_region) + 3
    ws2.cell(row=start, column=1, value="Revenue by Product").font = title_font
    _write_series(ws2, by_product, start_row=start + 1, header_font=header_font, header_fill=header_fill, body_font=body_font)

    start2 = start + 1 + len(by_product) + 3
    ws2.cell(row=start2, column=1, value="Revenue by Sales Rep").font = title_font
    _write_series(ws2, by_rep, start_row=start2 + 1, header_font=header_font, header_fill=header_fill, body_font=body_font)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    # --- Cleaned data sheet ---
    ws3 = wb.create_sheet("Cleaned Data")
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = body_font
    for col_cells in ws3.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws3.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 30)

    wb.save(REPORT_PATH)


def _write_series(ws, series, start_row, header_font, header_fill, body_font):
    ws.cell(row=start_row, column=1, value=series.index.name or "Category").font = header_font
    ws.cell(row=start_row, column=1).fill = header_fill
    ws.cell(row=start_row, column=2, value="Total Sale ($)").font = header_font
    ws.cell(row=start_row, column=2).fill = header_fill
    for i, (idx, val) in enumerate(series.items(), start=1):
        ws.cell(row=start_row + i, column=1, value=idx).font = body_font
        ws.cell(row=start_row + i, column=2, value=round(float(val), 2)).font = body_font


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    log_lines = []
    log("=== Data Cleaning & Reporting Automation ===", log_lines)

    df_raw = load_data(RAW_DATA_PATH)
    df_clean = clean_data(df_raw, log_lines)
    df_clean.to_csv(CLEANED_CSV_PATH, index=False)
    log(f"Cleaned data saved to {CLEANED_CSV_PATH}", log_lines)

    summary, by_region, by_product, by_rep, monthly = build_summary(df_clean)
    chart_paths = make_charts(by_region, by_product, by_rep, monthly)
    log(f"Charts saved to {CHART_DIR}/", log_lines)

    build_excel_report(df_clean, summary, by_region, by_product, by_rep, chart_paths)
    log(f"Excel report saved to {REPORT_PATH}", log_lines)

    CLEANING_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")
    log(f"Cleaning log saved to {CLEANING_LOG_PATH}", log_lines)
    log("=== Done ===", log_lines)


if __name__ == "__main__":
    main()
